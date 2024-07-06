from tkinter import *
from tkinter import ttk, messagebox, filedialog
import json
from datetime import datetime
from bs4 import BeautifulSoup
import requests
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import os
import urllib.parse
from unidecode import unidecode
from abc import ABC, abstractmethod
import ctypes
ctypes.windll.kernel32.FreeConsole()


class VagasCollector:
  def __init__(self):
    root.title('Vagas Collector')
    root.resizable(FALSE, FALSE)
    root.call('source', 'forest-dark.tcl')
    ttk.Style().theme_use('forest-dark')
    self.barraMenu(root)
    
    mainFrame = ttk.Frame(root, padding=15)
    mainFrame.grid(column=0, row=0)
    self.areaDeBusca(mainFrame)


  def barraMenu(self, root):
    root.option_add('*tearOff', FALSE)
    menubar = Menu(root)

    root['menu'] = menubar
    menu_arquivo = Menu(menubar)
    menubar.add_cascade(menu=menu_arquivo, label='Arquivo')
    menu_arquivo.add_command(label='Selecionar pasta...', command=self.selecionarDiretorio)

    menu_ajuda = Menu(menubar, name='help')
    menubar.add_cascade(menu=menu_ajuda, label='Ajuda')
    menu_ajuda.add_command(label ='Instruções', command=self.abrirInstrucoes) 
  

  def selecionarDiretorio(self):
    dirname = filedialog.askdirectory()
    if dirname:
      dados = {
        'escolha': True,
        'diretorio': dirname
      }
      with open ('settings.json', 'w') as arquivo:
        json.dump(dados, arquivo)
  

  def abrirInstrucoes(self):
    instrucoes = Toplevel(root)
    instrucoes.title('Instruções')
    instrucoes.resizable(FALSE, FALSE)
    instrucoes.attributes("-topmost", 1)
    instrucoes.grab_set()

    mainFrame = ttk.Frame(instrucoes, padding=(20, 20, 20, 30))
    mainFrame.grid(column=0, row=0)

    texto = ttk.Label(mainFrame, text="""Boas-vindas ao Vagas Collector! Para usar esta ferramenta, é muito simples:

  • Preencha o formulário de forma correta. Os únicos campos obrigatórios
    são o da vaga e a seleção de um dos portais de empregos.
  • Ao deixar o campo Quantidade em branco, o programa entenderá que você
    procura o máximo de vagas permitido, que é 250.
  • Para a busca no portal Indeed funcionar, você precisa ter instalado
    o navegador Google Chrome.
  • Ao fim da busca, você poderá salvar uma planilha com as ofertas encontradas.
    Basta escolher a pasta desejada para seu arquivo Excel. Para trocar a
    pasta, basta ir no menu e selecionar a opção presente na aba Arquivo.
  • O programa cria planilhas baseadas no nome da vaga procurada.
  • Se já existir uma planilha com o mesmo nome da vaga procurada na pasta
    desejada, o programa abre este arquivo e o atualiza com novas informações.

Você está pronto(a) para usar o programa sem problemas! Boa sorte em sua procura :)
                

Obs.: Este programa busca nos sites a depender do documento HTML dos mesmos. Ou
seja, provavelmente seu uso ficará obsoleto em algum momento, pois geralmente as
páginas na web são atualizadas com novas estruturas. Acontecendo isso, o programa
será atualizado para o bom funcionamento do mesmo.""")
    texto.grid(column=0, row=0)


  def areaDeBusca(self, parent):
    areaDeBusca = ttk.LabelFrame(parent, text='Área de busca', padding=10)
    areaDeBusca.grid(column=0, row=0)

    tituloDaVagaLabel = ttk.Frame(areaDeBusca)
    tituloDaVagaLabel.grid(column=0, row=0, columnspan=3, sticky=(W))

    asteriscoObrigatorio = ttk.Label(tituloDaVagaLabel, text="*", foreground='red')
    asteriscoObrigatorio.grid(column=0, row=0)

    tituloDaVaga = ttk.Label(tituloDaVagaLabel, text='Título da vaga:', padding=(0, 0, 0, 5))
    tituloDaVaga.grid(column=1, row=0)

    self.vaga = StringVar()
    vagaEntry = ttk.Entry(areaDeBusca, textvariable=self.vaga, width=40)
    vagaEntry.grid(column=0, row=1, columnspan=3)

    estadoEscolha = ttk.Label(areaDeBusca, text='Estado:', padding=(0, 10, 0, 5))
    estadoEscolha.grid(column=0, row=2, sticky=(W))

    self.estadoSelected = StringVar()

    estadoList = ttk.Combobox(areaDeBusca, textvariable=self.estadoSelected, state='readonly')
    estadoList['values'] = self.todosOsEstados()
    estadoList.grid(column=0, row=3, columnspan=2, sticky=(W))

    cidadeEscolha = ttk.Label(areaDeBusca, text='Cidade:', padding=(0, 10, 0, 5))
    cidadeEscolha.grid(column=0, row=4, sticky=(W))

    self.cidadeSelected = StringVar()

    cidadeList = ttk.Combobox(areaDeBusca, textvariable=self.cidadeSelected, state='readonly')
    cidadeList.configure(state='disabled')
    cidadeList.grid(column=0, row=5, columnspan=2, sticky=(W))

    quantidade = ttk.Label(areaDeBusca, text="Quantidade (max. 250):", padding=(0, 10, 0, 5))
    quantidade.grid(column=0, row=6, columnspan=2, sticky=(W))

    self.numeroQuantidade = StringVar()
    quantidadeEntry = ttk.Entry(areaDeBusca, textvariable=self.numeroQuantidade, width=10)
    quantidadeEntry.grid(column=0, row=7, sticky=(W))

    portaisFrame = ttk.Frame(areaDeBusca, padding=(0, 15, 0, 0))
    portaisFrame.grid(column=0, row=8, sticky=(W))

    self.botaoVagas = StringVar()
    checkVagas = ttk.Checkbutton(portaisFrame, text='Portal Vagas', variable=self.botaoVagas, onvalue='onVagas', offvalue='offVagas')
    checkVagas.grid(column=0, row=0)

    self.botaoIndeed = StringVar()
    checkIndeed = ttk.Checkbutton(portaisFrame, text='Portal Indeed', variable=self.botaoIndeed, onvalue='onIndeed', offvalue='offIndeed', padding=(0, 5, 0, 0))
    checkIndeed.grid(column=0, row=1)

    self.start = ttk.Button(areaDeBusca, text="Iniciar processo", style='Accent.TButton', command=self.iniciarProcesso)
    self.start.grid(column=1, row=8)

    estadoList.bind('<<ComboboxSelected>>', lambda e: self.todasAsCidades(cidadeList, self.estadoSelected.get()))


  def todosOsEstados(self):
    listaEstados = ['', 'Home Office']
    with open('locais.json', 'r', encoding='utf8') as arquivo:
      dados = json.load(arquivo)
    
    for estado in dados['estados']:
      listaEstados.append(estado['nome'])
    return listaEstados    
  

  def todasAsCidades(self, lista, valor):
    lista.set("")

    if valor in ['', 'Home Office']:
      lista.state(['disabled'])
    else:
      lista.state(['!disabled'])
      with open('locais.json', 'r', encoding='utf8') as arquivo:
        dados = json.load(arquivo)
      for estado in dados['estados']:
        if estado['nome'] == valor:
          listaCompleta = ['']
          for cidade in estado['cidades']:
            listaCompleta.append(cidade)
          lista['values'] = listaCompleta
          lista.state(['readonly'])
          break
  

  def iniciarProcesso(self):
    cargo = self.vaga.get()
    estado = self.estadoSelected.get()
    cidade = self.cidadeSelected.get()
    quantidade = self.numeroQuantidade.get()
    opcaoVagas = self.botaoVagas.get()
    opcaoIndeed = self.botaoIndeed.get()


    if len(cargo) > 30:
      messagebox.showwarning(message='Vaga não pode conter mais do que 30 caracteres!')
      return
    
    if cargo == "" or cargo.isspace():
      messagebox.showwarning(message='Vaga não pode ficar em branco!')
      return
    
    if not all(char.isalpha() or char.isspace() for char in cargo):
      messagebox.showwarning(message='Vaga deve conter apenas letras e espaços!')
      return
    
    try:
      quantidade = int(quantidade)
      if quantidade > 250:
        quantidade = 250
      if quantidade == 0:
        messagebox.showwarning(message='Quantidade não pode ser 0!')
        return
    except:
      if quantidade == "" or quantidade.isspace():
        quantidade = 250
      else:
        messagebox.showwarning(message='Quantidade só pode ficar em branco ou ler números!')
        return
    
    if opcaoVagas != 'onVagas' and opcaoIndeed != 'onIndeed':
      messagebox.showwarning(message='Escolha pelo menos um dos portais para iniciar o processo!')
      return
    
    self.start.configure(state='disabled')
    
    if estado == "":
      estado = False
    if cidade == "":
      cidade = False
    
    if opcaoVagas == 'onVagas':
      coletor = Vagas(
        cargo,
        estado,
        cidade,
        quantidade
      )
      self.coletarOportunidades(coletor, 'Vagas')
      

    if opcaoIndeed == 'onIndeed':
      coletor = Indeed(
        cargo,
        estado,
        cidade,
        quantidade
      )
      self.coletarOportunidades(coletor, 'Indeed')

    self.start.configure(state='!disabled')
  

  def coletarOportunidades(self, objeto, portal):
    objeto.procurarOfertas()
    vagas = objeto.coletarLista()
    if objeto.existeOferta:
      while True:
        for vaga in vagas:
          objeto.coletarOferta(vaga)
          if objeto.finish:
            break
        objeto.proximaPagina()
        vagas = objeto.coletarLista()
        if objeto.finish:
          break
      fazerPlanilha = messagebox.askyesno(message=f'No Portal {portal}, encontramos {len(objeto.oportunidades)} vaga(s) de {objeto.cargo} em {objeto.local} para você! Deseja salvar sua planilha?', icon='question', title='Resultado'
      )
      if fazerPlanilha:
        with open ('settings.json', 'r') as arquivo:
          dados = json.load(arquivo)
          if dados['escolha'] == False:
            dirname = filedialog.askdirectory()
            if dirname:
              dados = {
                  'escolha': True,
                  'diretorio': dirname
                }
              with open ('settings.json', 'w') as arquivo:
                json.dump(dados, arquivo)
              objeto.criarPlanilha(dirname)
              messagebox.showinfo(message=objeto.aviso)
            else:
              messagebox.showinfo(message='Sua planilha não foi criada.')
          else:
            diretorio = dados['diretorio']
            objeto.criarPlanilha(diretorio)
            messagebox.showinfo(message=objeto.aviso)
      else:
        messagebox.showinfo(message='Sua planilha não foi criada.')
    else:
      messagebox.showinfo(message=f'No Portal {portal}, não encontramos nenhuma vaga de {objeto.cargo} em {objeto.local} para você...')


def getSoup(url):
  request = requests.get(url)
  return BeautifulSoup(request.text, 'html.parser')


class Jobs(ABC):
  def __init__(self, cargo, estado=False, cidade=False, quantidade=250):
    self.cargo = cargo.strip()
    self.estado = estado
    self.cidade = cidade
    self.quantidade = quantidade
    self.oportunidades = []
    self.existeOferta = False
    self.finish = False
    self.erro = False
  

  @abstractmethod
  def procurarOfertas(self):
    pass
  

  @abstractmethod
  def coletarLista(self):
    pass


  @abstractmethod
  def coletarOferta(self, vaga):
    pass


  @abstractmethod
  def proximaPagina(self):
    pass


  @abstractmethod
  def saveOferta(self):
    pass
  

  def criarPlanilha(self, diretorio):
    
    dia = datetime.today().day
    mes = datetime.today().month
    ano = datetime.today().year

    plataforma = self.plataforma
    maxColunas = self.maxColunas

    self.data = f'{ano}_{mes}_{dia}'
    cargoFormated = self.cargo.split()
    self.cargoFormated = '_'.join(cargoFormated)
    self.arquivo = diretorio + f'/Portal_{plataforma}_{self.cargoFormated}.xlsx'
  
    if os.path.exists(self.arquivo):
      firstTime = False
      wb = load_workbook(filename=self.arquivo)
      ws = wb.active
    else:
      firstTime = True
      wb = Workbook()
      ws = wb.active
      ws.append(self.categorias)

    self.guardarDados(ws)

    if firstTime:
      tab = Table(displayName="TabelaVagas", ref=f"A1:{maxColunas}{len(self.oportunidades) + 1}")
      style = TableStyleInfo(name="TableStyleMedium18", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
      tab.tableStyleInfo = style
      ws.add_table(tab)

    wb.save(self.arquivo)
    self.oportunidades = []

    if firstTime: self.aviso = f'Planilha {self.plataforma.upper()} criada com sucesso.'
    else: self.aviso = f'Planilha {self.plataforma.upper()} atualizada com sucesso.'
  

  @abstractmethod
  def guardarDados(self, workSheet):
    pass


class OfertaVagas:
  def __init__(self, cargo, empresa, nivel, link, salario, local, contrato, beneficios, descricao):
    self.cargo = cargo
    self.empresa = empresa
    self.nivel = nivel
    self.link = link
    self.salario = salario
    self.local = local
    self.contrato = contrato
    self.beneficios = beneficios
    self.descricao = descricao


class OfertaIndeed:
  def __init__(self, cargo, empresa, local, link, salario, tipo, turno, beneficios, descricao):
    self.cargo = cargo
    self.empresa = empresa
    self.local = local
    self.link = link
    self.salario = salario
    self.tipo = tipo
    self.turno = turno
    self.beneficios = beneficios
    self.descricao = descricao



class Vagas(Jobs):
  def __init__(self, cargo, estado, cidade, quantidade):
    super().__init__(cargo, estado, cidade, quantidade)
    self.maxColunas = 'I'
    self.plataforma = 'Vagas'
    self.categorias = ['Cargo', 'Empresa', 'Nível', 'Link', 'Salário', 'Local', 'Contrato', 'Benefícios', 'Descrição']
    self.masterURL = 'https://www.vagas.com.br'
  

  def procurarOfertas(self):
    cargo = unidecode(self.cargo)
    cargo = cargo.lower()
    cargo = cargo.split()
    cargo = '-'.join(cargo)

    if self.cidade:
      estado = urllib.parse.quote(self.estado)
      cidade = unidecode(self.cidade)
      cidade = cidade.lower()
      cidade = cidade.split()
      cidade = '-'.join(cidade)
      self.soup = getSoup(self.masterURL + '/vagas-de-' + cargo + '-em-' + cidade + '?e%5B%5D=' + estado)
      self.local = self.estado + ', ' + self.cidade

    elif self.estado:
      self.local = self.estado

      if self.estado == 'Home Office':
        self.soup = getSoup(self.masterURL + '/vagas-de-' + cargo + '?m%5B%5D=100%25+Home+Office')
        return

      estado = urllib.parse.quote(self.estado)
      self.soup = getSoup(self.masterURL + '/vagas-de-' + cargo + '?e%5B%5D=' + estado)

    else:
      self.local = 'sua busca'
      self.soup = getSoup(self.masterURL + '/vagas-de-' + cargo)



  def coletarLista(self):
    if not self.finish:
      try:
        lista = self.soup.find(id='todasVagas')
        self.existeOferta = True
        return lista.find_all('li', 'vaga')
      except:
        self.aviso = f'Não encontramos vagas de {self.cargo} em {self.local}!'
        self.erro = True
        return
  

  def coletarOferta(self, vaga):
    cargo = vaga.find('h2', 'cargo').get_text().strip()
    empresa = vaga.find('span', 'emprVaga').get_text().strip()
    nivel = vaga.find('span', 'nivelVaga').get_text().strip()
    link = vaga.find('a')
    url = 'https://www.vagas.com.br' + link.get('href')

    self.soup2 = getSoup(url)
    infoVaga = self.soup2.find('div', 'infoVaga')
    infos = infoVaga.find_all('li')
    for i, info in enumerate(infos):
      if i == 0:
        spans = info.find_all('span')
        for span in spans:
          if span.get_text().strip() != 'Faixa salarial':
            salarioString = span.get_text().strip()
            salarioFormatado = salarioString.split()
            salario = ' '.join(salarioFormatado)
      
      elif i == 1:
        local = info.get_text().strip()
      
      else:
        contrato = info.get_text().strip()

    listaBeneficios = self.soup2.find_all('li', 'job-benefits__list-item')
    beneficiosList = []
    for beneficio in listaBeneficios:
      texto = beneficio.get_text()
      beneficiosList.append(texto.strip())
    beneficios = ', '.join(beneficiosList)
    if beneficios == "": beneficios = 'N/A'

    descricaoBox = self.soup2.find('div', 'job-description__text')
    descricao = descricaoBox.contents
    linhas = []
    linhasFormated = []
    for linha in descricao:
      linhas.append(linha.get_text().strip())

    for linha in linhas:
      if not linha.isspace() and linha not in ['Descrição', '', ':']:
        palavras = linha.split()
        novaLinhaLista = []

        for palavra in palavras:
          characters = list(palavra)
          novaPalavraLista = []

          for character in characters:
            if character != '﻿':
              novaPalavraLista.append(character)

          novaPalavra = ''.join(novaPalavraLista)
          novaLinhaLista.append(novaPalavra)

        linhaFormated = ' '.join(novaLinhaLista)
        linhasFormated.append(linhaFormated)

    descricao = ' '.join(linhasFormated)

    self.saveOferta(
      cargo, empresa, nivel,
      url, salario, local,
      contrato, beneficios, descricao
    )
  
    if len(self.oportunidades) == self.quantidade:
      self.finish = True
  

  def proximaPagina(self):
    if not self.finish:
      try:
        maisVagas = self.soup.find('a', 'btMaisVagas')
        url = maisVagas.get('data-url')
      except:
        self.finish = True
      else:
        self.soup = getSoup(self.masterURL + url)
  

  def saveOferta(self, cargo, empresa, nivel, link, salario, local, contrato, beneficios, descricao):
    oferta = OfertaVagas(
      cargo, empresa, nivel,
      link, salario, local,
      contrato, beneficios, descricao
    )
    self.oportunidades.append(oferta)
  

  def criarPlanilha(self, diretorio):
    return super().criarPlanilha(diretorio)
  

  def guardarDados(self, workSheet):
    for oportunidade in self.oportunidades:
      workSheet.append([
        oportunidade.cargo, oportunidade.empresa, oportunidade.nivel,
        oportunidade.link, oportunidade.salario, oportunidade.local,
        oportunidade.contrato, oportunidade.beneficios, oportunidade.descricao
      ])


class Indeed(Jobs):
  def __init__(self, cargo, estado, cidade, quantidade):
    super().__init__(cargo, estado, cidade, quantidade)
    self.maxColunas = 'I'
    self.plataforma = 'Indeed'
    self.categorias = ['Cargo', 'Empresa', 'Local', 'Link', 'Salário', 'Tipo de vaga', 'Turno/horário', 'Benefícios', 'Descrição']
  

  def procurarOfertas(self):
    cargo = self.cargo.split()
    cargo = '+'.join(cargo)

    self.options = Options()
    user_agent = 'Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/60.0.3112.50 Safari/537.36'
    self.options.add_argument(f'user-agent={user_agent}')
    self.options.add_argument("--headless")

    self.driver = webdriver.Chrome(options=self.options)
    self.driver.implicitly_wait(3)

    if self.cidade:
      self.local = self.estado + ', ' + self.cidade
      estado = self.estado.split()
      estado = '+'.join(estado)
      cidade = self.cidade.split()
      cidade  ='+'.join(cidade)
      self.driver.get('https://br.indeed.com/jobs?q=' + cargo + '&l=' + estado + '%2C+' + cidade)
    
    elif self.estado:
      self.local = self.estado

      if self.estado == 'Home Office':
        self.driver.get('https://br.indeed.com/jobs?q=' + cargo + '&l=remoto')
        return
      
      estado = self.estado.split()
      estado = '+'.join(estado)
      self.driver.get('https://br.indeed.com/jobs?q=' + cargo + '&l=' + estado)
    
    else:
      self.local = 'sua busca'
      self.driver.get('https://br.indeed.com/jobs?q=' + cargo)
  

  def coletarLista(self):
    if not self.finish:
      try:
        vagas = self.driver.find_elements(By.CLASS_NAME, 'job_seen_beacon')
      except:
        self.aviso = f'Não há vagas de {self.cargo} no {self.local}!'
        self.erro = True
        self.finish = True
      else:
        self.driverExtra = webdriver.Chrome(options=self.options)
        self.driverExtra.implicitly_wait(3)
        self.existeOferta = True
        return vagas

  def coletarOferta(self, vaga):

    link = vaga.find_element(By.CLASS_NAME, 'jcs-JobTitle')
    url = link.get_attribute('href')
    self.driverExtra.get(url)

    dadosBox = self.driverExtra.find_element(By.CLASS_NAME, 'jobsearch-InfoHeaderContainer')
    cargo = dadosBox.find_element(By.CLASS_NAME, 'jobsearch-JobInfoHeader-title').text.strip()
    empresaBox = dadosBox.find_element(By.CLASS_NAME, 'css-1h46us2')
    try:
      empresa = empresaBox.find_element(By.TAG_NAME, 'a').text.strip()
    except:
      empresa = empresaBox.find_element(By.CLASS_NAME, 'css-1z0pyms').text.strip()

    try:
      local = dadosBox.find_element(By.CLASS_NAME, 'css-waniwe').text.strip()
    except:
      local = dadosBox.find_element(By.CLASS_NAME, 'css-17cdm7w').text.strip()

    salario = "N/A"
    tipo = "N/A"
    turno = "N/A"

    try:
      dados = self.driverExtra.find_element(By.XPATH, '//*[@id="jobDetailsSection"]/div/div[1]/div[2]')
    except:
      pass
    else:
      divs = dados.find_elements(By.TAG_NAME, 'div')
      for div in divs:
        categoria = div.get_attribute('aria-label')

        if categoria == 'Salário':
          salario = div.find_element(By.CLASS_NAME, 'ecydgvn1').text.strip()

        elif categoria == 'Tipo de vaga':
          elementos = []
          lis = div.find_elements(By.TAG_NAME, 'li')
          for li in lis:
            elementos.append(li.text.strip())
          tipo = ', '.join(elementos)

        elif categoria == 'Turno e horário de trabalho':
          turno = div.find_element(By.TAG_NAME, 'li').text.strip()

    try:
      beneficiosBox = self.driverExtra.find_element(By.ID, 'benefits')
    except:
      beneficios = "N/A"
    else:
      lis = beneficiosBox.find_elements(By.TAG_NAME, 'li')
      elementos = []
      for li in lis:
        elementos.append(li.text.strip())
      beneficios = ', '.join(elementos)

    descricaoBox = self.driverExtra.find_element(By.ID, 'jobDescriptionText').text.strip()
    descricao = descricaoBox.split()
    descricaoFormated = ' '.join(descricao)

    self.saveOferta(
      cargo, empresa, local,
      url, salario, tipo,
      turno, beneficios, descricaoFormated
    )

    if len(self.oportunidades) == self.quantidade:
      self.finish = True
      self.driverExtra.quit()
      self.driver.quit()


  def proximaPagina(self):
    if not self.finish:
      try:
        navMenu = self.driver.find_element(By.CLASS_NAME, 'css-98e656')
        botoes = navMenu.find_elements(By.TAG_NAME, 'a')
        for botao in botoes:
          ariaLabel = botao.get_attribute('aria-label')
          if ariaLabel == 'Next Page':
            elemento = botao
        url = elemento.get_attribute('href')
      except:
        self.finish = True
        self.driverExtra.quit()
        self.driver.quit()
      else:
        self.driver.get(url)
  

  def saveOferta(self, cargo, empresa, local, link, salario, tipo, turno, beneficios, descricao):
    oferta = OfertaIndeed(
      cargo, empresa, local,
      link, salario, tipo,
      turno, beneficios, descricao
    )
    self.oportunidades.append(oferta)
  

  def criarPlanilha(self, diretorio):
    return super().criarPlanilha(diretorio)
  

  def guardarDados(self, workSheet):
    for oportunidade in self.oportunidades:
      workSheet.append([
        oportunidade.cargo, oportunidade.empresa, oportunidade.local,
        oportunidade.link, oportunidade.salario, oportunidade.tipo,
        oportunidade.turno, oportunidade.beneficios, oportunidade.descricao
      ])


root = Tk()
VagasCollector()
root.mainloop()